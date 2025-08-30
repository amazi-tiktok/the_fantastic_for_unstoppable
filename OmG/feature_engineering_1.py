#!/usr/bin/env python3
"""
Fixed Feature Engineering Script for Restaurant Reviews
Addresses the TF-IDF "empty vocabulary" error and provides comprehensive feature engineering
"""

import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report, accuracy_score
import re
import string
import warnings

warnings.filterwarnings('ignore')

try:
    import nltk

    # Download required NLTK data if needed
    try:
        nltk.data.find('tokenizers/punkt')
    except LookupError:
        nltk.download('punkt', quiet=True)
    try:
        nltk.data.find('corpora/stopwords')
    except LookupError:
        nltk.download('stopwords', quiet=True)
    from nltk.corpus import stopwords

    NLTK_AVAILABLE = True
except ImportError:
    print("NLTK not available, using basic text processing")
    NLTK_AVAILABLE = False


class RestaurantReviewFeatureEngineering:
    """
    Comprehensive feature engineering class for restaurant reviews
    """

    def __init__(self):
        self.tfidf_vectorizer = None
        self.label_encoder = None
        self.scaler = None
        self.feature_names = []

    def preprocess_text(self, text):
        """
        Comprehensive text preprocessing function
        """
        if pd.isna(text):
            return ""

        # Convert to lowercase
        text = str(text).lower()

        # Remove URLs
        text = re.sub(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', '', text)

        # Remove email addresses
        text = re.sub(r'\S+@\S+', '', text)

        # Remove extra whitespace and newlines
        text = re.sub(r'\s+', ' ', text)

        # Remove standalone numbers but keep words with numbers
        text = re.sub(r'\b\d+\b', '', text)

        # Remove excessive punctuation but keep sentence structure
        text = re.sub(r'[^\w\s.,!?;-]', '', text)

        # Remove extra spaces
        text = text.strip()

        return text

    def extract_text_features(self, df):
        """
        Extract various text-based features
        """
        # Basic text statistics
        df['text_length'] = df['cleaned_text'].str.len()
        df['word_count'] = df['cleaned_text'].str.split().str.len()
        df['sentence_count'] = df['cleaned_text'].str.count(r'[.!?]+') + 1
        df['avg_word_length'] = df['cleaned_text'].apply(
            lambda x: np.mean([len(word) for word in x.split()]) if x.split() else 0
        )

        # Punctuation features
        df['exclamation_count'] = df['cleaned_text'].str.count('!')
        df['question_count'] = df['cleaned_text'].str.count(r'\?')
        df['period_count'] = df['cleaned_text'].str.count(r'\.')
        df['uppercase_ratio'] = df['text'].apply(
            lambda x: sum(1 for c in str(x) if c.isupper()) / len(str(x)) if len(str(x)) > 0 else 0
        )

        # Sentiment indicators (simple word-based)
        positive_words = ['good', 'great', 'excellent', 'amazing', 'wonderful', 'fantastic',
                          'love', 'perfect', 'delicious', 'tasty', 'nice', 'beautiful',
                          'clean', 'fresh', 'recommend', 'best', 'awesome', 'incredible']
        negative_words = ['bad', 'terrible', 'awful', 'horrible', 'disgusting', 'dirty',
                          'expensive', 'slow', 'cold', 'stale', 'rotten', 'disappointed',
                          'worst', 'hate', 'poor', 'terrible']

        df['positive_word_count'] = df['cleaned_text'].apply(
            lambda x: sum([1 for word in positive_words if word in x.lower()])
        )
        df['negative_word_count'] = df['cleaned_text'].apply(
            lambda x: sum([1 for word in negative_words if word in x.lower()])
        )
        df['sentiment_ratio'] = df['positive_word_count'] / (df['negative_word_count'] + 1)

        # Restaurant-specific features
        food_words = ['food', 'taste', 'flavor', 'delicious', 'meal', 'dish', 'menu', 'eat']
        service_words = ['service', 'staff', 'waiter', 'waitress', 'employee', 'server']
        atmosphere_words = ['atmosphere', 'ambiance', 'environment', 'place', 'location', 'view']
        price_words = ['price', 'expensive', 'cheap', 'cost', 'affordable', 'reasonable']

        df['food_mentions'] = df['cleaned_text'].apply(
            lambda x: sum([1 for word in food_words if word in x.lower()])
        )
        df['service_mentions'] = df['cleaned_text'].apply(
            lambda x: sum([1 for word in service_words if word in x.lower()])
        )
        df['atmosphere_mentions'] = df['cleaned_text'].apply(
            lambda x: sum([1 for word in atmosphere_words if word in x.lower()])
        )
        df['price_mentions'] = df['cleaned_text'].apply(
            lambda x: sum([1 for word in price_words if word in x.lower()])
        )

        return df

    def safe_tfidf_transform(self, texts, max_features=1000, min_df=2, max_df=0.95):
        """
        Safe TF-IDF transformation with error handling and progressive fallback
        """
        # Remove any empty texts
        texts = [text for text in texts if text and len(text.strip()) > 0]

        if len(texts) == 0:
            raise ValueError("No valid texts found for TF-IDF transformation")

        print(f"Attempting TF-IDF with {len(texts)} documents...")

        # Progressive fallback strategy
        strategies = [
            {
                'max_features': min(max_features, len(texts) // 2),
                'min_df': max(2, min(min_df, len(texts) // 10)),
                'max_df': max_df,
                'ngram_range': (1, 2),
                'stop_words': 'english',
                'token_pattern': r'\b[a-zA-Z]{2,}\b'
            },
            {
                'max_features': min(500, len(texts) // 3),
                'min_df': max(1, len(texts) // 20),
                'max_df': 0.98,
                'ngram_range': (1, 1),
                'stop_words': 'english',
                'token_pattern': r'\b[a-zA-Z]{2,}\b'
            },
            {
                'max_features': min(300, len(texts) // 5),
                'min_df': 1,
                'max_df': 0.99,
                'ngram_range': (1, 1),
                'stop_words': None,
                'token_pattern': r'\b\w+\b'
            }
        ]

        for i, strategy in enumerate(strategies):
            try:
                print(f"Trying strategy {i + 1}: max_features={strategy['max_features']}, "
                      f"min_df={strategy['min_df']}, max_df={strategy['max_df']}")

                tfidf = TfidfVectorizer(
                    max_features=strategy['max_features'],
                    min_df=strategy['min_df'],
                    max_df=strategy['max_df'],
                    ngram_range=strategy['ngram_range'],
                    stop_words=strategy['stop_words'],
                    lowercase=True,
                    strip_accents='unicode',
                    token_pattern=strategy['token_pattern']
                )

                tfidf_matrix = tfidf.fit_transform(texts)

                print(f"✓ TF-IDF successful! Shape: {tfidf_matrix.shape}")
                print(f"✓ Vocabulary size: {len(tfidf.vocabulary_)}")

                self.tfidf_vectorizer = tfidf
                return tfidf_matrix, tfidf

            except ValueError as e:
                print(f"✗ Strategy {i + 1} failed: {e}")
                continue

        # If all strategies fail, use Count Vectorizer as final fallback
        print("All TF-IDF strategies failed, trying Count Vectorizer...")
        try:
            count_vectorizer = CountVectorizer(
                max_features=min(200, len(texts) // 5),
                min_df=1,
                max_df=0.99,
                ngram_range=(1, 1),
                stop_words=None,
                lowercase=True,
                token_pattern=r'\b\w+\b'
            )
            count_matrix = count_vectorizer.fit_transform(texts)
            print(f"✓ Count Vectorizer successful! Shape: {count_matrix.shape}")
            self.tfidf_vectorizer = count_vectorizer
            return count_matrix, count_vectorizer
        except Exception as final_e:
            print(f"✗ Final fallback failed: {final_e}")
            return None, None

    def encode_categorical_features(self, df):
        """
        Encode categorical features properly
        """
        # Label encode the target variable (rating)
        self.label_encoder = LabelEncoder()
        df['rating_encoded'] = self.label_encoder.fit_transform(df['rating'])

        # Group low-frequency business names
        business_counts = df['business_name'].value_counts()
        top_businesses = business_counts.head(15).index  # Top 15 businesses
        df['business_name_grouped'] = df['business_name'].apply(
            lambda x: x if x in top_businesses else 'Other'
        )

        # One-hot encode categorical features
        categorical_features = ['business_name_grouped', 'rating_category']
        encoded_features = []

        for feature in categorical_features:
            if feature in df.columns:
                dummies = pd.get_dummies(df[feature], prefix=feature, drop_first=True)
                encoded_features.append(dummies)
                print(f"Encoded {feature}: {dummies.shape[1]} features")

        if encoded_features:
            categorical_encoded = pd.concat(encoded_features, axis=1)
            return categorical_encoded
        else:
            return pd.DataFrame()

    def fit_transform(self, df):
        """
        Complete feature engineering pipeline
        """
        print("=== Starting Feature Engineering Pipeline ===")

        # 1. Text preprocessing
        print("\n1. Preprocessing text...")
        df['cleaned_text'] = df['text'].apply(self.preprocess_text)

        # Remove completely empty texts
        empty_mask = df['cleaned_text'].str.len() > 0
        df = df[empty_mask].reset_index(drop=True)
        print(f"   Kept {len(df)} samples after removing empty texts")

        # 2. Extract text features
        print("\n2. Extracting text features...")
        df = self.extract_text_features(df)

        # 3. TF-IDF transformation
        print("\n3. Applying TF-IDF transformation...")
        tfidf_matrix, vectorizer = self.safe_tfidf_transform(df['cleaned_text'].tolist())

        if tfidf_matrix is not None:
            tfidf_feature_names = [f'tfidf_{i}' for i in range(tfidf_matrix.shape[1])]
            tfidf_df = pd.DataFrame(tfidf_matrix.toarray(), columns=tfidf_feature_names)
        else:
            print("   Warning: TF-IDF failed, proceeding without text vectorization")
            tfidf_df = pd.DataFrame()

        # 4. Encode categorical features
        print("\n4. Encoding categorical features...")
        categorical_encoded = self.encode_categorical_features(df)

        # 5. Combine all features
        print("\n5. Combining all features...")

        # Select numerical text features
        numerical_features = [
            'text_length', 'word_count', 'sentence_count', 'avg_word_length',
            'exclamation_count', 'question_count', 'period_count', 'uppercase_ratio',
            'positive_word_count', 'negative_word_count', 'sentiment_ratio',
            'food_mentions', 'service_mentions', 'atmosphere_mentions', 'price_mentions'
        ]

        feature_components = []
        feature_names = []

        # Add numerical features
        if all(col in df.columns for col in numerical_features):
            numerical_df = df[numerical_features]
            feature_components.append(numerical_df)
            feature_names.extend(numerical_features)
            print(f"   Added {len(numerical_features)} numerical features")

        # Add TF-IDF features
        if not tfidf_df.empty:
            feature_components.append(tfidf_df)
            feature_names.extend(tfidf_df.columns.tolist())
            print(f"   Added {tfidf_df.shape[1]} TF-IDF features")

        # Add categorical features
        if not categorical_encoded.empty:
            feature_components.append(categorical_encoded)
            feature_names.extend(categorical_encoded.columns.tolist())
            print(f"   Added {categorical_encoded.shape[1]} categorical features")

        # Combine all features
        if feature_components:
            X = pd.concat(feature_components, axis=1)
            y = df['rating_encoded']
            self.feature_names = feature_names

            print(f"\n=== Feature Engineering Complete ===")
            print(f"Final feature matrix shape: {X.shape}")
            print(f"Target variable shape: {y.shape}")
            print(f"Missing values: {X.isnull().sum().sum()}")

            return X, y, df
        else:
            raise ValueError("No features could be created!")

    def get_feature_importance(self, X, y, n_estimators=100):
        """
        Train a model and return feature importance
        """
        # Split data
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.2, random_state=42, stratify=y
        )

        # Scale features
        self.scaler = StandardScaler()
        X_train_scaled = self.scaler.fit_transform(X_train)
        X_test_scaled = self.scaler.transform(X_test)

        # Train model
        rf_model = RandomForestClassifier(
            n_estimators=n_estimators,
            random_state=42,
            max_depth=10,
            min_samples_split=5,
            min_samples_leaf=2
        )
        rf_model.fit(X_train_scaled, y_train)

        # Predictions
        y_pred = rf_model.predict(X_test_scaled)
        accuracy = accuracy_score(y_test, y_pred)

        # Feature importance
        feature_importance = pd.DataFrame({
            'feature': self.feature_names,
            'importance': rf_model.feature_importances_
        }).sort_values('importance', ascending=False)

        return {
            'model': rf_model,
            'accuracy': accuracy,
            'feature_importance': feature_importance,
            'predictions': y_pred,
            'y_test': y_test
        }


import os
import json
import pandas as pd
from openpyxl import load_workbook

JSON_DIR = r"C:\Users\Om Gorakhia\OneDrive\Desktop\Om Gorakhia\Hackathon\Tiktok\category_reviews\category_reviews"
OUTPUT_FILE = 'processed_reviews_2.xlsx'
BATCH_SIZE = 4000

def get_processed_user_ids():
    if not os.path.exists(OUTPUT_FILE):
        return set()

    try:
        existing_df = pd.read_excel(OUTPUT_FILE)
        return set(existing_df['user_id'].astype(str))
    except Exception:
        return set()

import re

# Function to clean illegal Excel characters from strings
def clean_illegal_chars(value):
    if isinstance(value, str):
        # Remove control characters except \n, \r, and \t
        return re.sub(r"[\x00-\x08\x0b-\x0c\x0e-\x1f\x7f]", "", value)
    return value

# Updated append function with cleaning step
def append_to_excel(batch_df):
    # Clean all string columns to remove illegal characters
    for col in batch_df.select_dtypes(include=['object']).columns:
        batch_df[col] = batch_df[col].apply(clean_illegal_chars)

    if not os.path.exists(OUTPUT_FILE):
        batch_df.to_excel(OUTPUT_FILE, index=False)
    else:
        with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            book = load_workbook(OUTPUT_FILE)
            sheet = book.active
            start_row = sheet.max_row
            batch_df.to_excel(writer, index=False, header=False, startrow=start_row)


def parse_json_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
        if isinstance(data, dict):  # Single entry
            return [data]
        elif isinstance(data, list):  # List of entries
            return data
        return []

def main():
    all_files = [f for f in os.listdir(JSON_DIR) if f.endswith('.json')]
    processed_user_ids = get_processed_user_ids()
    batch = []

    print(f"Found {len(all_files)} JSON files to process...")

    for file_name in all_files:
        file_path = os.path.join(JSON_DIR, file_name)
        print(f"Processing file: {file_name}")

        records = parse_json_file(file_path)
        for record in records:
            if record['user_id'] in processed_user_ids:
                continue

            batch.append(record)
            processed_user_ids.add(record['user_id'])

            if len(batch) >= BATCH_SIZE:
                batch_df = pd.DataFrame(batch)
                append_to_excel(batch_df)
                print(f"Saved batch of {len(batch)} records to Excel.")
                batch = []

    if batch:  # Save remaining records
        batch_df = pd.DataFrame(batch)
        append_to_excel(batch_df)
        print(f"Saved final batch of {len(batch)} records to Excel.")

    print("All files processed.")

if __name__ == "__main__":
    main()
